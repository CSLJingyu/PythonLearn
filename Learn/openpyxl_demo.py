# 对于excel的数据进行读写
import openpyxl
import requests
import re


# 获取html文件
def get_html():
    url = 'https://www.weather.com.cn/weather1d/101010100.shtml'
    # 打开浏览器并且打开网址
    response = requests.get(url)

    # 设置编码格式
    response.encoding = 'utf-8'

    return response.text


# 解析html文件
def parse_html(response_text):
    # 正则表达式 爬取指定的内容
    city = re.findall('<span class="name">([\u4e00-\u9fa5]*)</span>', response_text)
    weather = re.findall('<span class="weather">([\u4e00-\u9fa5]*)</span>', response_text)
    lst = []
    for a, b in zip(city, weather):
        lst.append([a, b])

    return lst


html = get_html()
lst = parse_html(html)

# 写入数据
# 创建一个excel工作溥
workbook = openpyxl.Workbook()
# 创建工作表
sheet = workbook.create_sheet('景区天气')
# 给工作表中添加数据：
for item in lst:
    sheet.append(item)

workbook.save('C:\\Users\\舒言\\Desktop\\x.xlsx')

# 读取数据
# # 打开工作溥
workbook = openpyxl.load_workbook("景区天气.xlsx")

# # 获取工作表
sheet = workbook['景区天气']

# 读取单元格数据
# 表格数据二维列表 先行后列
# 存储行数据
lst_row = []
for row in sheet.rows:
    # 存储一行中全部的单元格数据
    sub_lst_row = []
    # 遍历每一行中全部单元格数据
    for cell in row:
        sub_lst_row.append(cell)

    # 存储遍历过的行
    lst_row.append(sub_lst_row)


# 测试存储的数据是否正确
for item in lst:
    print(item)
